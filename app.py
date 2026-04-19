import streamlit as st
import pandas as pd
from datetime import datetime
import sqlite3
from io import BytesIO

st.set_page_config(page_title="Le Nettoyeur Excel", page_icon="🧹", layout="centered")

st.title("🧹 Le Nettoyeur Excel & Migrateur")
st.markdown("Transformez vos fichiers Excel brouillons en données propres et prêtes pour une base de données sécurisée.")

# 1. THE UPLOAD ZONE
uploaded_file = st.file_uploader("Glissez votre fichier Excel ici (.xlsx)", type=["xlsx"])

if uploaded_file:
    # Read the raw, messy data
    df_raw = pd.read_excel(uploaded_file)
    
    st.subheader("🔍 Phase 1 : Audit des Données (Avant)")
    st.dataframe(df_raw.head(5))
    
    # Calculate errors
    duplicates = df_raw.duplicated().sum()
    missing_values = df_raw.isnull().sum().sum()
    
    col1, col2, col3 = st.columns(3)
    col1.metric("Lignes Totales", len(df_raw))
    col2.metric("Doublons Détectés", duplicates, delta_color="inverse")
    col3.metric("Cellules Vides", missing_values, delta_color="inverse")

    if st.button("🚀 Lancer le Nettoyage et la Migration"):
        with st.spinner("Nettoyage en cours..."):
            
            df_clean = df_raw.copy()
            
            # --- THE PANDAS WASHING MACHINE ---
            # --- THE PANDAS WASHING MACHINE ---
            # 1. Drop completely empty rows
            df_clean = df_clean.dropna(how='all')
            
            # 2. Remove exact duplicates
            df_clean = df_clean.drop_duplicates()
            
            # 3. 🧠 ADVANCED NUMBER EXTRACTION (Fixes "150 UNITS" & "ON QUOTE")
            for col in df_clean.columns:
                # Target columns that should be math (Prices, Quantities, Stock)
                if "PRIX" in col.upper() or "QUANTIT" in col.upper() or "STOCK" in col.upper():
                    # Use Regex to extract only digits and decimals from the string
                    extracted_numbers = df_clean[col].astype(str).str.extract(r'(\d+\.?\d*)')[0]
                    # Convert to math numbers. Things like "ON QUOTE" become NaN, which we turn to 0
                    df_clean[col] = pd.to_numeric(extracted_numbers, errors='coerce').fillna(0)
            
            # 4. 🧠 L'IMPUTATION INTELLIGENTE (Smart Data Recovery)
            
            # Step A: Standardize all text and convert "fake" blanks to real Pandas NA
            for col in df_clean.select_dtypes(include=['object']).columns:
                df_clean[col] = df_clean[col].astype(str).str.strip().str.upper()
                df_clean[col] = df_clean[col].replace(['NAN', '', 'NAT', 'UNKNOWN_DATE', 'DATE_INCONNUE'], pd.NA)

            # Step B: Relational Filling (Deduce Supplier & City based on Product history!)
            if 'Désignation_Produit' in df_clean.columns:
                if 'Fournisseur_Principal' in df_clean.columns:
                    # Group by product, find the most frequent supplier, and fill the blanks!
                    df_clean['Fournisseur_Principal'] = df_clean.groupby('Désignation_Produit')['Fournisseur_Principal'].transform(
                        lambda x: x.fillna(x.mode()[0] if not x.mode().empty else pd.NA)
                    )
                
                if 'Ville_Dépôt' in df_clean.columns:
                    df_clean['Ville_Dépôt'] = df_clean.groupby('Désignation_Produit')['Ville_Dépôt'].transform(
                        lambda x: x.fillna(x.mode()[0] if not x.mode().empty else pd.NA)
                    )

            # Step C: Smart Fallbacks for IDs and Dates
            for col in df_clean.columns:
                # 1. Rescue missing Reference IDs
                if "REF" in col.upper() or "ID" in col.upper():
                    mask = df_clean[col].isna()
                    df_clean.loc[mask, col] = [f"AUTO-RECUP-{i}" for i in range(1, mask.sum() + 1)]
                
                # 2. Rescue missing Dates (Use today's date)
                elif "DATE" in col.upper():
                    today_str = datetime.now().strftime("%Y-%m-%d")
                    df_clean[col] = df_clean[col].fillna(today_str)
                
                # 3. The Final Safety Net (For anything that couldn't be deduced)
                elif df_clean[col].dtype == 'object':
                    df_clean[col] = df_clean[col].fillna('INCONNU')
            
            # --- THE MIGRATION (TO SQLITE) ---
            # Connect to a local SQLite database file
            db_name = "factory_archive.db"
            conn = sqlite3.connect(db_name)
            # Write the clean dataframe to a SQL table
            df_clean.to_sql("clean_inventory", conn, if_exists="replace", index=False)
            conn.close()

            st.success("✅ Nettoyage terminé et migré avec succès vers la base de données (factory_archive.db) !")
            
            st.subheader("✨ Phase 2 : Données Propres (Après)")
            st.dataframe(df_clean.head(5))
            
            # --- THE OPENPYXL MAGIC (Styling the Excel File) ---
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_clean.to_excel(writer, index=False, sheet_name="Data_Propre")
                
                # Accéder au fichier Excel en arrière-plan
                workbook = writer.book
                worksheet = writer.sheets["Data_Propre"]
                
                from openpyxl.styles import PatternFill, Font, Alignment
                from openpyxl.utils import get_column_letter
                
                # 1. Designer l'en-tête (Fond bleu foncé, texte blanc)
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

                # 2. Ajustement automatique des colonnes et formatage
                for i, col in enumerate(df_clean.columns):
                    col_letter = get_column_letter(i + 1)
                    
                    # 🛡️ THE FIX: Utilisation sécurisée de .str.len() native à Pandas
                    max_data_len = df_clean[col].astype(str).str.len().max()
                    # Si la colonne est complètement vide, max() renvoie NaN. On le force à 0.
                    max_data_len = 0 if pd.isna(max_data_len) else int(max_data_len)
                    
                    max_length = max(max_data_len, len(str(col))) + 2
                    worksheet.column_dimensions[col_letter].width = max_length
                    
                    # Parcourir les lignes pour les couleurs et la monnaie
                    for row in range(2, len(df_clean) + 2):
                        cell = worksheet[f"{col_letter}{row}"]
                        
                        # Ajouter le format "MAD" pour la colonne de prix
                        if "PRIX" in col.upper():
                            cell.number_format = '#,##0.00 "MAD"'
                        
                        # Alerte Couleur : Rouge si stock < 20, Vert si > 100
                        if col == "Quantite_Stock":
                            try:
                                val = float(cell.value)
                                if val < 20:
                                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                    cell.font = Font(color="9C0006")
                                elif val > 100:
                                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                    cell.font = Font(color="006100")
                            except:
                                pass
            
            # --- LES BOUTONS DE TÉLÉCHARGEMENT ---
            col_btn1, col_btn2 = st.columns(2)
            
            with col_btn1:
                st.download_button(
                    label="📥 Télécharger l'Excel Stylisé",
                    data=output.getvalue(),
                    file_name="inventaire_sidi_ghanem_propre.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            with col_btn2:
                # Récupérer le fichier SQLite local pour l'offrir en téléchargement direct
                with open(db_name, "rb") as db_file:
                    st.download_button(
                        label="🗄️ Télécharger la Base SQL (.db)",
                        data=db_file,
                        file_name=db_name,
                        mime="application/x-sqlite3",
                        use_container_width=True
                    )
            
            st.info("💡 Conseil : Le fichier Excel est pour vous. Le fichier .db est la version sécurisée que je peux brancher sur un tableau de bord Power BI ou Azure.")